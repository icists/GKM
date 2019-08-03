import json
from openpyxl import Workbook

# for excel
write_wb = Workbook()
write_ws = write_wb.active

total = 0
user_num = 0 # application
apply_user_num = 0 # 200자 이상 작성자 + HIT + sm-pair + OB + BIZ-World
kaist_num = 0 #
foreign_num = 0 # 
male_num = 0 # 
passed_num = 0 
cancel_num = 0



essay1_1 = 0
essay1_2 = 0
essay2_1 = 0
essay2_2 = 0
dorm_num = 0
friend_num = 0


# bool
passed_ok = False
cancel_ok = False

with open('ams-icists-182409-export.json', encoding='utf-8') as json_file:
    json_data = json.load(json_file)
    users = json_data["users"]

    # for excel
    write_ws.cell(row=1, column=1).value = "email"
    write_ws.cell(row=1, column=2).value = "FirstName"
    write_ws.cell(row=1, column=3).value = "LastName"
    write_ws.cell(row=1, column=4).value = "essayWordCount"
    write_ws.cell(row=1, column=5).value = "groupName"
    write_ws.cell(row=1, column=6).value = "nationality"
    write_ws.cell(row=1, column=7).value = "school"
    write_ws.cell(row=1, column=8).value = "sex"
    write_ws.cell(row=1, column=9).value = "phoneNumber"    
    write_ws.cell(row=1, column=10).value = "essayTopic"
    write_ws.cell(row=1, column=11).value = "essay"
    write_ws.cell(row=1, column=12).value = "birthDate"
    write_ws.cell(row=1, column=13).value = "지원경로"
    write_ws.cell(row=1, column=14).value = "기숙사 사용"
    write_ws.cell(row=1, column=15).value = "financialAid"
    write_ws.cell(row=1, column=16).value = "financialAidEssay"
    write_ws.cell(row=1, column=17).value = "major"
    write_ws.cell(row=1, column=18).value = "prevParticipation"
    write_ws.cell(row=1, column=19).value = "visaSupport"
    write_ws.cell(row=1, column=20).value = "otherChannel"
    write_ws.cell(row=1, column=21).value = "lastUpdate"
    write_ws.cell(row=1, column=22).value = "passed"
    write_ws.cell(row=1, column=23).value = "cancel"


    for i in users:
        passed_ok = False
        cancel_ok = False

        total += 1

        user = users[i]

        if "email" in user:
            email = user["email"]
            write_ws.cell(row=total + 1, column=1).value = email
            if email == 'koha1003@naver.com':
                user_num += 1
                passed_num += 1
                essay2_1 += 1
                friend_num += 1
        if "application" in user:
            user_num += 1

            # var
            application = user["application"]
            FirstName = application["nameFirst"]
            LastName = application["nameLast"]
            essayWordCount = application["essayWordCount"]
            groupName = application["groupName"]
            nationality = application["nationality"]
            school = application["school"]
            sex = application["sex"]
            phoneNumber = application["phoneNumber"]
            essayTopic = application["essayTopic"]
            essay = application["essay"]
            birthDate = application["birthDate"]
            channel = application["channel"]
            dormUse = application["dormUse"]
            financialAid = application["financialAid"]
            financialAidEssay = application["financialAidEssay"]
            major = application["major"]
            prevParticipation = application["prevParticipation"]
            visaSupport = application["visaSupport"]
            otherChannel = application["otherChannel"]
            lastUpdate = application["lastUpdate"]

            if "passed" in application:
                passed = application["passed"]
                passed_ok = True

            if "cancel" in application:
                cancel = application["cancel"]
                cancel_ok = True

            str(phoneNumber)
            if '+82' in phoneNumber:
                # +821012345678
                if len(phoneNumber) == 13:
                    phoneNumber = phoneNumber.replace('+82', '0')
                elif len(phoneNumber) == 12:
                    phoneNumber = phoneNumber.replace('+82', '01')

            # for write excel
            write_ws.cell(row=total + 1, column=2).value = FirstName
            write_ws.cell(row=total + 1, column=3).value = LastName
            write_ws.cell(row=total + 1, column=4).value = essayWordCount
            write_ws.cell(row=total + 1, column=5).value = groupName
            write_ws.cell(row=total + 1, column=6).value = nationality
            write_ws.cell(row=total + 1, column=7).value = school
            write_ws.cell(row=total + 1, column=8).value = sex
            write_ws.cell(row=total + 1, column=9).value = phoneNumber
            write_ws.cell(row=total + 1, column=10).value = essayTopic
            write_ws.cell(row=total + 1, column=11).value = essay
            write_ws.cell(row=total + 1, column=12).value = birthDate
            write_ws.cell(row=total + 1, column=13).value = channel
            write_ws.cell(row=total + 1, column=14).value = dormUse
            write_ws.cell(row=total + 1, column=15).value = financialAid
            write_ws.cell(row=total + 1, column=16).value = financialAidEssay
            write_ws.cell(row=total + 1, column=17).value = major
            write_ws.cell(row=total + 1, column=18).value = prevParticipation
            write_ws.cell(row=total + 1, column=19).value = visaSupport
            write_ws.cell(row=total + 1, column=20).value = otherChannel
            write_ws.cell(row=total + 1, column=21).value = lastUpdate
            if passed_ok:
                write_ws.cell(row=total + 1, column=22).value = passed
            if cancel_ok:
                write_ws.cell(row=total + 1, column=23).value = cancel


            #count
            if essayWordCount > 200 or groupName.upper() == 'HIT' or groupName.upper() == 'SM-PAIR' or 'BIZWORLD' in groupName.upper() or email == '1220859289@qq.com' or email == 'yshin0917@gmail.com' or email == 'blessedita@kaist.ac.kr':
                apply_user_num += 1

                if sex == 'Male': male_num += 1
                if school.upper() == 'KAIST': kaist_num += 1
                if nationality != 'Korea, Republic of': foreign_num += 1

                if passed_ok and passed:
                    passed_num += 1
                if cancel_ok and cancel:
                    cancel_num += 1

                if essayTopic == 'essay1_1':
                    essay1_1 += 1
                elif essayTopic == 'essay1_2':
                    essay1_2 += 1
                elif essayTopic == 'essay2_1':
                    essay2_1 += 1
                else:
                    essay2_2 += 1

                if dormUse:
                    dorm_num += 1

                if channel == 'Friend':
                    friend_num += 1




#save excel
write_wb.save('2019ICISTS.xlsx')

#출력
print('----------------통계----------------')
print('가입자 수: %s' %total)
print('지원서 작성 중: %s' %user_num)

print('')
print('지원서 작성 완료 - 취소 수: %s' %(apply_user_num - cancel_num))
print('얼리 합격자: %s' %passed_num)
print('레귤러 지원: %s' %(apply_user_num - passed_num - cancel_num))
print('취소자 수: %s' %cancel_num)
print('')
print('----------------지원자 분류---------------')
print('KAIST 학생: %s' %kaist_num)
print('타대생: %s' %(apply_user_num-kaist_num))
print('외국인: %s' %foreign_num)
print('남성: %s' %male_num)
print('여성: %s' %(apply_user_num-male_num))
print('')
print('----------------기타 정보---------------')
print('에세이 1-1: %d 1-2: %d 2-1: %d 2-2: %d' %(essay1_1, essay1_2, essay2_1, essay2_2))
print('지인 홍보: %s' %friend_num)