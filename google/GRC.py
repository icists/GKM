import openpyxl
import math

# Start
print("----------start----------\n")

# definition
def reward_count(team_number, cost):
  for money in enter_value[team_number]:
    person_reward[money] += total_list[team_number][money] * cost
    chk_regal[money] += total_list[team_number][money]

def reward_end(total, cnt, end = 0):
  for i in range(total):
    team_number = top_enter[i][1]
    if i < cnt:
      reward_count(team_number, 0.193 + end)
    else:
      reward_count(team_number, 0.097 + end)

def enter_rank(num):
  rank = 1
  for i in top_enter:
    if rank <= num:
      rank += 1
      continue
    print("Rank #%s: %s" %(rank, total_list[i[1]][0]))
    rank += 1

def check_regal(person, enter_num, bound):
  sum = 0
  for i in range(2, enter_num + 2):
    sum += total_list[person][i]
    if(sum > bound):
      print("%s(이)가 금액을 초과했습니다." %person_enter[i])


def make_line():
  print('')

# Variables
round_number = int(input('Round number ?: '))
print('')

total_list = [[0] * 300 for i in range(300)]
enter_value = [0 for i in range(300)]
person_enter = [0 for i in range(300)]
belong_enter = [0 for i in range(300)]  
person_reward = [0 for i in range(300)]
chk_regal = [0 for i in range(300)]
top_enter = []
enter_choice = []
col_num = 0
row_num = 0
enter_name = ''
enter_sum = 0
person_num = 0
enter_num = 0

# Open Excel
filename = input('.xlsx 포함한 입력 파일명: ')
grc_wb = openpyxl.load_workbook(filename)
grc = grc_wb.worksheets[0]
print('')

# Read Excel

# Calculate Sigma
for col in grc.columns:

  # Person info
  if col_num < 2:


    for people in col:
      total_list[col_num][row_num] = people.value
      
      if row_num == 0: 
        row_num += 1
        continue
      
      if col_num == 0:
        person_enter[row_num] = people.value
      else:
        belong_enter[row_num] = people.value

      row_num += 1

    person_num = row_num
    col_num += 1
    row_num = 0

    continue
  
  # Sum Value
  for enter in col:
    total_list[col_num][row_num] = enter.value
    if row_num == 0:
      enter_name = str(enter.value)
      row_num += 1
      continue

    if enter.value != 0:
      enter_choice.append(row_num)

    enter_sum += int(enter.value)

    row_num += 1
  
  top_enter.append((enter_sum, col_num))
  enter_value[col_num] = enter_choice
  print(enter_name + ' 총합: ' + str(enter_sum))

  # Reset
  chk = True
  col_num += 1
  row_num = 0
  enter_sum = 0
  enter_choice = []

enter_num = col_num - 2

# Sorting
top_enter.sort(reverse=True)

# Check Round
if round_number == 1:
  reward_end(8, 5)
elif round_number == 2:
  reward_end(5, 3)
elif round_number == 3:
  reward_end(2, 1)
else:
  reward_end(2, 1, 1)

# Print
make_line()

for cnt in range(1, person_num):
  print("%s(%s): %.2f" %(person_enter[cnt], belong_enter[cnt], person_reward[cnt]))

make_line()

print('----------탈락----------')
if round_number == 1:
  enter_rank(10)
elif round_number == 2:
  enter_rank(6)
elif round_number == 3:
  enter_rank(3)
else:
  enter_rank(3)

# Write Excel
grc_wb = openpyxl.Workbook()
grc_ws = grc_wb.active

for c in range(3):
  for l in range(person_num):
    if c == 2:
      if l == 0: grc_ws.cell(row=l+1, column=c+1).value = '리워드'
      else: grc_ws.cell(row=l+1, column=c+1).value = person_reward[l]
    else:
      grc_ws.cell(row=l+1, column=c+1).value = total_list[c][l]

grc_wb.save('GRC.xlsx')

# Done
done = input('\nPress Enter to Exit... \n')
print('----------Done----------')